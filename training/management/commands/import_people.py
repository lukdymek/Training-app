from django.core.management.base import BaseCommand
from django.db import transaction
from training.models import Person
from openpyxl import load_workbook
from datetime import datetime, date


def normalize_header(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "_")


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    # try common text formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except Exception:
            pass
    return None  # if unknown format


class Command(BaseCommand):
    help = "Import People from an Opera Excel export (xlsx). Updates/creates by SYSPer ID. Optionally archives missing."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument("--sheet", type=str, default=None, help="Sheet name (optional)")
        parser.add_argument("--dry-run", action="store_true", help="Parse but do not write to DB")

        # Default behavior: archive missing people (recommended)
        parser.add_argument(
            "--archive-missing",
            action="store_true",
            default=True,
            help="Archive people not present in the file (sets is_active=False). Default: ON",
        )
        parser.add_argument(
            "--no-archive-missing",
            dest="archive_missing",
            action="store_false",
            help="Do NOT archive people missing from the file.",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        xlsx_path = options["xlsx_path"]
        sheet_name = options["sheet"]
        dry_run = options["dry_run"]
        archive_missing = options["archive_missing"]

        wb = load_workbook(filename=xlsx_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        # Read header row (first row)
        headers = [normalize_header(cell.value) for cell in ws[1]]

        col = {h: i for i, h in enumerate(headers)}

        def pick(*names):
            for n in names:
                if n in col:
                    return col[n]
            return None

        idx_sysper = pick("sysper_id", "sysperid", "sysper", "id")
        idx_first = pick("first_name", "firstname", "first", "name", "given_name")
        idx_last = pick("last_name", "lastname", "last", "surname", "family_name")
        idx_email = pick("email", "email_address", "e-mail")
        idx_dob = pick("dob", "date_of_birth", "birth_date")
        idx_gender = pick("gender", "sex")
        idx_category = pick("category", "employee_category", "cat")
        idx_deploy = pick("current_deployment", "deployment", "current_assignment", "deployed")

        if idx_sysper is None:
            raise Exception(f"Could not find SYSPer column in headers. Found headers: {headers}")

        created = 0
        updated = 0
        skipped = 0

        incoming_ids = set()

        # Loop through rows starting at row 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            sysper_val = row[idx_sysper]
            if sysper_val is None or str(sysper_val).strip() == "":
                skipped += 1
                continue

            try:
                sysper_id = int(str(sysper_val).strip())
            except ValueError:
                skipped += 1
                continue

            incoming_ids.add(sysper_id)

            first_name = (row[idx_first] if idx_first is not None else "") or ""
            last_name = (row[idx_last] if idx_last is not None else "") or ""
            email = (row[idx_email] if idx_email is not None else "") or ""
            dob = parse_date(row[idx_dob] if idx_dob is not None else None)
            gender = (row[idx_gender] if idx_gender is not None else "") or ""
            category = (row[idx_category] if idx_category is not None else "") or ""
            current_deployment = (row[idx_deploy] if idx_deploy is not None else "") or ""

            obj, was_created = Person.objects.update_or_create(
                sysper_id=sysper_id,
                defaults={
                    "first_name": str(first_name).strip(),
                    "last_name": str(last_name).strip(),
                    "email": str(email).strip(),
                    "dob": dob,
                    "gender": str(gender).strip(),
                    "category": str(category).strip(),
                    "current_deployment": str(current_deployment).strip(),
                    "is_active": True,  # ✅ active if present in file
                },
            )

            if was_created:
                created += 1
            else:
                updated += 1

        archived = 0

        # Archive people not present in the file (recommended: keeps history intact)
        if (not dry_run) and archive_missing:
            archived = Person.objects.exclude(sysper_id__in=incoming_ids).update(is_active=False)

        if dry_run:
            # Roll back everything in dry-run (including updates/creates)
            transaction.set_rollback(True)
            archived = 0  # dry-run doesn't actually archive

        self.stdout.write(self.style.SUCCESS(
            f"Import complete. created={created}, updated={updated}, archived={archived}, skipped={skipped}, "
            f"dry_run={dry_run}, archive_missing={archive_missing}"
        ))
