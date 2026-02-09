
from .models import Person, Training, Participation, Subject, TrainerSkill
from django.contrib import admin, messages
from django.urls import path
from django.shortcuts import render, redirect
from django import forms
from django.db.models import Q
from django.views.decorators.http import require_http_methods
from django.utils.decorators import method_decorator
from openpyxl import load_workbook
from datetime import datetime, date
from django.db import transaction
from django.utils.dateparse import parse_date
import os
import tempfile
from .models import EmailTemplate, EmailRecipient, EmailRecipientGroup




@admin.register(Subject)
class SubjectAdmin(admin.ModelAdmin):
    list_display = ("name", "is_recurring", "validity_days")
    list_editable = ("is_recurring", "validity_days")
    search_fields = ("name",)  # allows autocomplete + searching subjects


class TrainerSkillInline(admin.TabularInline):
    model = TrainerSkill
    extra = 0
    autocomplete_fields = ("subject",)





@admin.register(Training)
class TrainingAdmin(admin.ModelAdmin):
    list_display = ("course_name", "start_at", "end_at", "location", "capacity")
    search_fields = ("course_name", "location")
    ordering = ("-start_at",)


@admin.register(Participation)
class ParticipationAdmin(admin.ModelAdmin):
    list_display = ("person", "training", "role")
    search_fields = ("person__sysper_id", "person__last_name", "training__course_name")
    list_filter = ("role",)


@admin.register(TrainerSkill)
class TrainerSkillAdmin(admin.ModelAdmin):
    list_display = ("trainer", "subject")
    list_select_related = ("trainer", "subject")
    search_fields = (
        "subject__name",
        "trainer__first_name",
        "trainer__last_name",
        "trainer__sysper_id",
    )
    list_filter = ("subject",)
    search_help_text = "Search by subject name, trainer name, or SYSPER ID"
    autocomplete_fields = ("trainer", "subject")


def normalize_header(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "_")


def parse_date(value):
    if value is None or value == "":
        return None

    # IMPORTANT: datetime must be checked BEFORE date
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    s = str(value).strip()

    # Handle ISO strings like 1900-01-01 or 1900-01-01T00:00:00
    try:
        # Python can parse both date and datetime ISO formats
        dt = datetime.fromisoformat(s)
        return dt.date()
    except Exception:
        pass

    try:
        return date.fromisoformat(s)
    except Exception:
        pass

    # Try common text formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    return None


class PeopleImportForm(forms.Form):
    file = forms.FileField(help_text="Upload Opera Evo Excel export (.xlsx)")
    replace_missing = forms.BooleanField(
        required=False,
        initial=True,
        help_text="Archive people not present in this file (recommended)."
    )


@admin.register(Person)
class PersonAdmin(admin.ModelAdmin):
    list_display = ("sysper_id", "last_name", "first_name", "email", "category", "current_deployment", "is_active")
    ordering = ("last_name", "first_name")
    search_fields = ("sysper_id", "last_name", "first_name", "email")
    inlines = [TrainerSkillInline]

    # adds a button on the changelist page
    change_list_template = "admin/training/person_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "import-excel/",
                self.admin_site.admin_view(self.import_people_excel),
                name="training_person_import_excel",
            ),
        ]
        return custom + urls

    @transaction.atomic
    def import_people_excel(self, request):
        """
        Step 1 (POST without confirm):
            Upload Excel -> parse -> preview (store file path in session)

        Step 2 (POST with confirm=1):
            Re-read Excel from temp file -> bulk upsert to DB
            Optionally archive missing people (is_active=False)
        """

        # =====================================================
        # Helper: read workbook and parse rows from it
        # =====================================================
        def parse_people_from_workbook(workbook):
            ws = workbook.active

            headers = [normalize_header(c.value) for c in ws[1]]
            col = {h: i for i, h in enumerate(headers)}

            def pick(*names):
                for n in names:
                    if n in col:
                        return col[n]
                return None

            idx_sysper = pick("sysper_id", "sysperid", "sysper", "id")
            idx_first = pick("first_name", "firstname", "first", "name", "given_name")
            idx_last = pick("last_name", "lastname", "last", "surname", "family_name")
            idx_email = pick("email", "email_address", "e-mail")
            idx_dob = pick("dob", "date_of_birth", "birth_date")
            idx_gender = pick("gender", "sex")
            idx_category = pick("category", "employee_category", "cat")
            idx_deploy = pick("current_deployment", "deployment", "current_assignment", "deployed")

            if idx_sysper is None:
                raise ValueError(f"Could not find SYSPer ID column. Headers detected: {headers}")

            rows = []
            incoming_ids = set()
            skipped = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                sysper_val = row[idx_sysper]
                if not sysper_val or str(sysper_val).strip() == "":
                    skipped += 1
                    continue

                try:
                    sysper_id = int(str(sysper_val).strip())
                except ValueError:
                    skipped += 1
                    continue

                incoming_ids.add(sysper_id)

                first_name = (row[idx_first] if idx_first is not None else "") or ""
                last_name = (row[idx_last] if idx_last is not None else "") or ""
                email = (row[idx_email] if idx_email is not None else "") or ""

                dob_date = parse_date(row[idx_dob] if idx_dob is not None else None)
                # Keep DOB as ISO string if we need preview; for DB we convert back to date
                dob_iso = dob_date.isoformat() if dob_date else ""

                gender = (row[idx_gender] if idx_gender is not None else "") or ""
                category = (row[idx_category] if idx_category is not None else "") or ""
                current_deployment = (row[idx_deploy] if idx_deploy is not None else "") or ""

                rows.append({
                    "sysper_id": sysper_id,
                    "first_name": str(first_name).strip(),
                    "last_name": str(last_name).strip(),
                    "email": str(email).strip(),
                    "dob": dob_iso,
                    "gender": str(gender).strip(),
                    "category": str(category).strip(),
                    "current_deployment": str(current_deployment).strip(),
                })

            return headers, rows, incoming_ids, skipped


        # =====================================================
        # GET — show upload form
        # =====================================================
        if request.method == "GET":
            return render(
                request,
                "admin/training/import_people.html",
                {"form": PeopleImportForm()},
            )

        # =====================================================
        # POST — CONFIRM IMPORT
        # =====================================================
        if request.POST.get("confirm") == "1":
            tmp_path = request.session.get("people_import_tmp_path")
            archive_missing = request.session.get("people_import_archive_missing", True)

            if not tmp_path or not os.path.exists(tmp_path):
                messages.error(request, "No pending import file found. Please upload the file again.")
                return redirect("..")

            try:
                wb = load_workbook(filename=tmp_path, data_only=True)
                headers, rows, incoming_ids, skipped = parse_people_from_workbook(wb)
            except Exception as e:
                messages.error(request, f"Failed to read the uploaded file again: {e}")
                return redirect("..")

            # De-duplicate by sysper_id (if file has duplicates, last row wins)
            rows_by_id = {}
            for r in rows:
                sid = r.get("sysper_id")
                if sid:
                    rows_by_id[int(sid)] = r

            incoming_ids = set(rows_by_id.keys())

            # Counts (optional, but nice to show)
            existing_ids = set(
                Person.objects.filter(sysper_id__in=incoming_ids)
                .values_list("sysper_id", flat=True)
            )
            created = len(incoming_ids - existing_ids)
            updated = len(incoming_ids & existing_ids)

            # Build model objects for bulk upsert
            objs = []
            for sysper_id, r in rows_by_id.items():
                dob = parse_date(r.get("dob"))  # convert ISO string back to date
                objs.append(
                    Person(
                        sysper_id=sysper_id,
                        first_name=(r.get("first_name") or "").strip(),
                        last_name=(r.get("last_name") or "").strip(),
                        email=(r.get("email") or "").strip(),
                        dob=dob,
                        gender=(r.get("gender") or "").strip(),
                        category=(r.get("category") or "").strip(),
                        current_deployment=(r.get("current_deployment") or "").strip(),
                        is_active=True,
                    )
                )

            # IMPORTANT: fast DB write (batched UPSERT)
            with transaction.atomic():
                Person.objects.bulk_create(
                    objs,
                    update_conflicts=True,          # enables "upsert" on Postgres
                    unique_fields=["sysper_id"],    # conflict target
                    update_fields=[
                        "first_name",
                        "last_name",
                        "email",
                        "dob",
                        "gender",
                        "category",
                        "current_deployment",
                        "is_active",
                    ],
                    batch_size=1000,                # tune if needed
                )

                archived = 0
                if archive_missing:
                    archived = Person.objects.exclude(sysper_id__in=incoming_ids).update(is_active=False)

            # Cleanup: remove temp file and session keys
            try:
                os.remove(tmp_path)
            except Exception:
                pass

            request.session.pop("people_import_tmp_path", None)
            request.session.pop("people_import_archive_missing", None)

            messages.success(
                request,
                f"Import complete. created={created}, updated={updated}, archived={archived}, skipped={skipped}",
            )
            return redirect("..")

        # =====================================================
        # POST — UPLOAD / PREVIEW STEP
        # =====================================================

        form = PeopleImportForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, "admin/training/import_people.html", {"form": form})

        f = form.cleaned_data["file"]

        # Your form field is called replace_missing, but old code used archive_missing.
        # We'll support both safely:
        archive_missing = form.cleaned_data.get("replace_missing", True)

        # Save uploaded file to a temp file (so we don't store thousands of rows in session)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        try:
            for chunk in f.chunks():
                tmp.write(chunk)
            tmp.flush()
            tmp_path = tmp.name
        finally:
            tmp.close()

        try:
            wb = load_workbook(filename=tmp_path, data_only=True)
            headers, rows, incoming_ids, skipped = parse_people_from_workbook(wb)
        except Exception as e:
            try:
                os.remove(tmp_path)
            except Exception:
                pass
            messages.error(request, f"Could not read this Excel file: {e}")
            return render(request, "admin/training/import_people.html", {"form": PeopleImportForm()})

        existing_ids = set(
            Person.objects.filter(sysper_id__in=incoming_ids)
            .values_list("sysper_id", flat=True)
        )

        will_update = sum(1 for r in rows if r["sysper_id"] in existing_ids)
        will_create = len(rows) - will_update
        will_archive = (
            Person.objects.exclude(sysper_id__in=incoming_ids).count()
            if archive_missing
            else 0
        )

        # Store only temp file path + settings in session
        request.session["people_import_tmp_path"] = tmp_path
        request.session["people_import_archive_missing"] = archive_missing

        return render(
            request,
            "admin/training/import_people_preview.html",
            {
                "count": len(rows),
                "skipped": skipped,
                "preview": rows[:10],
                "will_create": will_create,
                "will_update": will_update,
                "will_archive": will_archive,
                "archive_missing": archive_missing,
                "headers": headers,
            },
        )


@admin.register(EmailTemplate)
class EmailTemplateAdmin(admin.ModelAdmin):
    list_display = ("name", "kind", "is_active", "updated_at")
    list_filter = ("kind", "is_active")
    search_fields = ("name", "subject", "body")


@admin.register(EmailRecipient)
class EmailRecipientAdmin(admin.ModelAdmin):
    list_display = ("name", "email", "is_active")
    list_filter = ("is_active",)
    search_fields = ("name", "email")
    ordering = ("name",)

@admin.register(EmailRecipientGroup)
class EmailRecipientGroupAdmin(admin.ModelAdmin):
    list_display = ("name", "is_active")
    list_filter = ("is_active",)
    search_fields = ("name",)
    filter_horizontal = ("recipients",)  # this MUST be the M2M field name